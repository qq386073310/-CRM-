import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from typing import List, Dict, Optional, Set
from datetime import datetime
from core.logger import logger

class ImportExportError(Exception):
    """Base exception for import/export operations"""
    pass

class DuplicateDataHandler:
    """Mixin class for handling duplicate data"""
    
    def __init__(self):
        self._existing_data_cache: Set[str] = set()
    
    def cache_existing_data(self, data: List[Dict[str, str]], key_columns: List[str]) -> None:
        """Cache existing data for duplicate checking"""
        self._existing_data_cache.clear()
        for row in data:
            key = self._generate_row_key(row, key_columns)
            self._existing_data_cache.add(key)
    
    def is_duplicate(self, row: Dict[str, str], key_columns: List[str]) -> bool:
        """Check if row is duplicate based on key columns"""
        key = self._generate_row_key(row, key_columns)
        return key in self._existing_data_cache
    
    @staticmethod
    def _generate_row_key(row: Dict[str, str], key_columns: List[str]) -> str:
        """Generate a unique key for the row based on key columns"""
        return '|'.join(str(row.get(col, '')) for col in key_columns)

class BaseImporterExporter(DuplicateDataHandler):
    """Base class for import/export operations"""
    
    def __init__(self):
        super().__init__()
        self._required_columns: List[str] = []
        self._column_mapping: Dict[str, str] = {}
    
    def export_to_excel(self, 
                       data: List[Dict[str, str]], 
                       file_path: str, 
                       sheet_name: str = 'Data',
                       headers: Optional[List[str]] = None) -> None:
        """
        Export data to Excel file
        
        Args:
            data: List of dictionaries containing the data
            file_path: Path to save the Excel file
            sheet_name: Name of the worksheet
            headers: Optional list of column headers to use
        """
        logger.info(f"\n=== 开始Excel导出 ===")
        logger.info(f"目标文件: {file_path}")
        logger.info(f"数据记录数: {len(data)}")
        
        if not data:
            raise ImportExportError("No data to export")
            
        # 打印前3条数据用于调试
        logger.debug("\n前3条数据内容:")
        for i, row in enumerate(data[:3], 1):
            logger.debug(f"记录 {i}: {row}")
            
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Use provided headers or infer from first data row
        headers = headers or list(data[0].keys())
        logger.debug(f"\n使用的表头: {headers}")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data with detailed logging
        logger.debug("\n写入数据过程:")
        for row_num, row in enumerate(data, 2):
            if row_num <= 4:  # 打印前3行写入过程
                logger.debug(f"正在写入行 {row_num}:")
            for col_num, header in enumerate(headers, 1):
                value = row.get(header, '')
                if row_num <= 4:  # 打印前3行每个单元格的值
                    logger.debug(f"  列 {col_num}({header}): {value}")
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        try:
            logger.info(f"正在保存文件到: {file_path}")
            
            wb.save(file_path)
            logger.info("文件保存成功")
            
            # 验证文件是否生成
            if not os.path.exists(file_path):
                raise ImportExportError("文件保存失败，未找到生成的文件")
                
            # 验证文件大小
            file_size = os.path.getsize(file_path)
            logger.info(f"文件大小: {file_size} 字节")
            
            # 重新打开验证内容
            verify_wb = load_workbook(file_path)
            verify_sheet = verify_wb.active
            logger.info(f"工作表行数: {verify_sheet.max_row}")
            logger.info(f"工作表列数: {verify_sheet.max_column}")
            
            # 验证数据行数
            if verify_sheet.max_row < 2:  # 只有表头
                logger.warning("警告: 导出的文件似乎为空(只有表头)")
                    
        except Exception as e:
            logger.error(f"文件保存失败: {str(e)}")
            raise ImportExportError(f"Failed to save Excel file: {str(e)}")
    
    def import_from_excel(self, 
                         file_path: str, 
                         sheet_name: str = 'Data',
                         key_columns: Optional[List[str]] = None,
                         skip_duplicates: bool = True) -> List[Dict[str, str]]:
        """
        Import data from Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet to read
            key_columns: Columns to use for duplicate checking
            skip_duplicates: Whether to skip duplicate rows
            
        Returns:
            List of dictionaries containing the imported data
        """
        if not os.path.exists(file_path):
            raise ImportExportError("File not found")
            
        try:
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb[sheet_name]
        except Exception as e:
            raise ImportExportError(f"Failed to load Excel file: {str(e)}")
        
        # Read headers
        headers = [cell.value for cell in ws[1]]
        
        # Validate required columns
        missing_columns = [col for col in self._required_columns if col not in headers]
        if missing_columns:
            raise ImportExportError(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Read data
        imported_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            
            # Apply column mapping if defined
            if self._column_mapping:
                row_data = {self._column_mapping.get(k, k): v for k, v in row_data.items()}
            
            # Skip empty rows
            if not any(row_data.values()):
                continue
                
            # Skip duplicates if enabled
            if skip_duplicates and key_columns and self.is_duplicate(row_data, key_columns):
                continue
                
            imported_data.append(row_data)
        
        return imported_data
    
    def validate_data(self, data: Dict[str, str]) -> bool:
        """Validate data before import/export"""
        # Base implementation does basic validation
        # Should be overridden by subclasses for specific validation
        return True